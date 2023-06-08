#! Python3
# Takes a text file and converts it to an excel spreadsheet.
# All contents will be added to the next available column.
# One text file per column, one line of text per row.

import openpyxl, sys
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active


for column in range(1, sheet.max_column+1):
    newTextFile = open(f'textFromExcel{column}.txt', 'w')
    for row in range(1, sheet.max_row+1):
        currentCell = get_column_letter(column) + str(row)
        if sheet[currentCell].value == None:
            newTextFile.write("\n")
        else:
            newTextFile.write(sheet[currentCell].value)
    newTextFile.close()
        


        

