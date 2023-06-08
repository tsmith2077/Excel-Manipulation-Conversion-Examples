#! python3
# Takes a number N and creates a multiplication
# table in an excel sheet.
# Ex: python3 multiplicationTable.py N
# N = the highest multiplier in the table

import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

N = int(sys.argv[1])

# Row 1 and column A should be in Bold
for num1 in range(1, N+1):
    # Labels for multiplication table
    columnALabels = 'A' + str(num1+1)
    row1Labels = get_column_letter(int(num1+1)) + '1'
    sheet[columnALabels].font = Font(bold=True)
    sheet[row1Labels].font = Font(bold=True)
    sheet[columnALabels].value = num1
    sheet[row1Labels].value = num1
    for num2 in range(1, N+1):
        currentCell = get_column_letter(num2+1) + str(num1+1)
        sheet[currentCell].value = num1 * num2

wb.save('multiplicationTable.xlsx')