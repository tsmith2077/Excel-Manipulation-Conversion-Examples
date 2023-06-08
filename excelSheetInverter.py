#! Python3
# Inverts columns on excel spreadsheet.
# Ex: Row 3 Column 1 becomes Row 1, Column 3

import openpyxl, sys

fileToAdjust = sys.argv[1]

wb = openpyxl.load_workbook(fileToAdjust)
sheet = wb.active

# sheetData[rowNum][columnNum] used to access data from cell.
sheetData = []

for rowNum in range(1, (sheet.max_row+1)):
    rowList = []
    for columnNum in range(1, (sheet.max_column+1)):
        currentCellValue = sheet.cell(row=rowNum, column=columnNum).value
        rowList.append(currentCellValue)
    sheetData.append(rowList)

sheet.delete_cols(1,sheet.max_column)
sheet.delete_rows(1, sheet.max_row)

# Invert the row and column numbers
for rowNum in range(0, len(sheetData)):
    for columnNum in range(0, len(sheetData[rowNum])):
        sheet.cell(row=(columnNum+1), column=(rowNum+1)).value = sheetData[rowNum][columnNum]

wb.save('copy_' + fileToAdjust)