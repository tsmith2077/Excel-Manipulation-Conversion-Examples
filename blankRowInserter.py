#! Python3
# Inserts blank rows in excel document.
# Ex: python3 blankRowInserter.py N M file.xlsx
# N = row you'd like to insert rows before
# M = number of rows to be inserted
# file.xlsx file name you'd like to adjust

import openpyxl, sys

insertRowsBefore = int(sys.argv[1])
numBlankRowsToAdd = int(sys.argv[2])
fileToAdjust = sys.argv[3]

wb = openpyxl.load_workbook(fileToAdjust)
sheet = wb.active

for rowNum in range(1, (sheet.max_row+1)):
    for columnNum in range(1, (sheet.max_column+1)):
        if rowNum == insertRowsBefore and columnNum==1:
            sheet.insert_rows(idx=insertRowsBefore, amount=numBlankRowsToAdd)
        currentCellValue = sheet.cell(row=rowNum, column=columnNum).value
        sheet.cell(row=rowNum, column=columnNum).value = currentCellValue
        

wb.save('copy_' + fileToAdjust)