import openpyxl, sys
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook('')
sheet = wb.active

for textFile in range(1, len(sys.argv)):
    text = open(sys.argv[textFile])
    print(text)
    for line in range(0, len(text)):
        currentCell = get_column_letter(textFile) + str(line+1)
        sheet[currentCell].value = text[line]
    text.close()
        
wb.save('textToExcel.xlsx')